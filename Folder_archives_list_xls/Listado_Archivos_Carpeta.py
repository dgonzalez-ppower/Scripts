import os
import sys
import argparse
import platform
import datetime
import pandas as pd
from tkinter import Tk, Label, Button, Checkbutton, IntVar, filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink

# Función para saber si un archivo es oculto
def esta_oculto(filepath):
    if platform.system() == 'Windows':
        import ctypes
        attribute = ctypes.windll.kernel32.GetFileAttributesW(str(filepath))
        return attribute != -1 and (attribute & 2)
    else:
        return os.path.basename(filepath).startswith('.')

# Filtro para excluir archivos basura
def es_archivo_valido(filepath):
    nombre = os.path.basename(filepath).lower()
    extensiones_excluidas = ('.tmp', '.log', '.bak', '.temp', '.ds_store', 'thumbs.db', 'desktop.ini')
    return (
        not nombre.startswith('~') and
        not nombre.endswith(extensiones_excluidas) and
        not esta_oculto(filepath) and
        os.path.isfile(filepath) and
        os.path.getsize(filepath) > 0
    )

# Función principal para escanear carpeta
def escanear_carpeta(carpeta, recursivo):
    archivos = []
    for root_dir, _, files in os.walk(carpeta):
        for file in files:
            ruta_completa = os.path.join(root_dir, file)
            if es_archivo_valido(ruta_completa):
                archivos.append({
                    "Archivo": file,
                    "Ruta completa": ruta_completa,
                    "Tamaño (KB)": round(os.path.getsize(ruta_completa) / 1024, 2)
                })
        if not recursivo:
            break
    return archivos

# Guardar a Excel con hipervínculos y autofiltro
def exportar_excel(carpeta, archivos):
    df = pd.DataFrame(archivos)
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    nombre_excel = os.path.join(carpeta, f"listado_archivos_{timestamp}.xlsx")
    df.to_excel(nombre_excel, index=False)

    wb = load_workbook(nombre_excel)
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        ruta = ws.cell(row=row, column=2).value
        nombre = ws.cell(row=row, column=1).value
        ws.cell(row=row, column=1).value = f'=HYPERLINK("{ruta}", "{nombre}")'

    ws.auto_filter.ref = ws.dimensions
    wb.save(nombre_excel)
    return nombre_excel

# GUI
def lanzar_gui():
    def ejecutar():
        carpeta = filedialog.askdirectory(title="Selecciona la carpeta")
        if not carpeta:
            return
        recursivo = bool(recursivo_var.get())
        archivos = escanear_carpeta(carpeta, recursivo)
        if not archivos:
            messagebox.showwarning("Vacío", "No se encontraron archivos válidos.")
            return
        salida = exportar_excel(carpeta, archivos)
        messagebox.showinfo("¡Éxito!", f"Listado guardado como:\n{salida}")

    gui = Tk()
    gui.title("Listado de Archivos")

    Label(gui, text="¿Buscar de forma recursiva?").grid(row=0, column=0, padx=10, pady=10)
    recursivo_var = IntVar()
    Checkbutton(gui, variable=recursivo_var).grid(row=0, column=1)

    Button(gui, text="Seleccionar carpeta y generar", command=ejecutar).grid(row=1, column=0, columnspan=2, padx=10, pady=10)

    gui.mainloop()

# CLI
def lanzar_cli(args):
    if not os.path.isdir(args.folder):
        print(f"[ERROR] Ruta no válida: {args.folder}")
        sys.exit(1)
    archivos = escanear_carpeta(args.folder, args.recursivo)
    if not archivos:
        print("[INFO] No se encontraron archivos válidos.")
        sys.exit(0)
    salida = exportar_excel(args.folder, archivos)
    print(f"[OK] Listado generado: {salida}")

# MAIN
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Genera un listado de archivos con hipervínculos en Excel.")
    parser.add_argument("folder", nargs="?", help="Ruta de la carpeta a escanear")
    parser.add_argument("-r", "--recursivo", action="store_true", help="Buscar de forma recursiva")

    args = parser.parse_args()

    if args.folder:
        lanzar_cli(args)
    else:
        lanzar_gui()
