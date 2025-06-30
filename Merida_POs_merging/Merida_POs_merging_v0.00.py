import os
import re
import unicodedata
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Constants ---
VALID_EXTENSIONS = ['.xls', '.xlsx', '.xlsm']
HEADER_PATTERNS = [("C36", "J36"), ("C37", "J37"), ("B36", "I36"), ("B37", "I37")]
STANDARD_HEADERS = ["No", "CODIGO IMP", "DESCRIPCIÓN", "UNIDAD", "CANTIDAD", "P. UNITARIO", "MONTO"]
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# --- GUI ---
def choose_folder(prompt):
	root = tk.Tk()
	root.withdraw()
	return filedialog.askdirectory(title=prompt)

# --- Utilities ---
def normalize_header(text):
	text = re.sub(r'\s+', ' ', str(text).replace('\n', ' ').replace('\r', '')).strip().upper()
	if 'COD' in text:
		return "CODIGO IMP"
	elif 'DES' in text:
		return "DESCRIPCIÓN"
	elif 'UNI' in text:
		return "UNIDAD"
	elif 'CANT' in text:
		return "CANTIDAD"
	elif 'P' in text and 'UNIT' in text:
		return "P. UNITARIO"
	elif 'IMP' in text or 'MONTO' in text:
		return "MONTO"
	elif 'NO' in text:
		return "No"
	return text

def strip_accents(text):
	return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

def find_excel_files(root_folder):
	target_files = []
	for dirpath, _, files in os.walk(root_folder):
		for file in files:
			if any(file.lower().endswith(ext) for ext in VALID_EXTENSIONS):
				name = re.sub(r'[^\w\s]', '', file.upper())
				if name.startswith("PMXCCC"):
					target_files.append(os.path.join(dirpath, file))
	return target_files

# --- Main Logic ---
def extract_po_data(filepath, log):
	try:
		wb = load_workbook(filepath, data_only=True)
		ws = wb.active
		po_number = os.path.splitext(os.path.basename(filepath))[0]

		# Detect revision by scanning rows 1 and 2
		revision = "NO DATA FOUND"
		for row in [ws[1], ws[2]]:
			for cell in row:
				if cell.value and isinstance(cell.value, str) and "rev" in cell.value.lower():
					revision = cell.value.strip()
					break
				if revision != "NO DATA FOUND":
					break

		# Detect supplier name by scanning for "razon" and "social" in range A1:K20
		supplier = "NO DATA FOUND"
		for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=11):
			for idx, cell in enumerate(row):
				if cell.value and isinstance(cell.value, str):
					val = strip_accents(cell.value.lower())
					if "razon" in val and "social" in val:
						# Look right from current cell
						for right_cell in row[idx+1:]:
							if right_cell.value and isinstance(right_cell.value, str):
								supplier = right_cell.value.strip()
								break
						break
				if supplier != "NO DATA FOUND":
					break
			if supplier != "NO DATA FOUND":
				break

		headers = []

		# Locate headers
		for start, end in HEADER_PATTERNS:
			try:
				range_cells = ws[start:end]
				headers = [normalize_header(cell.value) for cell in range_cells[0]]
				if any(h in headers for h in ["CODIGO IMP", "COD"]):
					break
			except:
				continue
		if not headers:
			log.append(f"{po_number}: Headers not found")
			return []

		# Read items
		data_rows = []
		blank_counter = 0
		row_num = 37
		while True:
			row = ws[row_num]
			ef_values = [row[4].value, row[5].value]  # Columns E and F (0-based)
			col_h_val = row[7].value  # Column H

			# Check subtotal condition
			if col_h_val and isinstance(col_h_val, str) and "SUBTOTAL" in col_h_val.upper():
				break

			# Check if row is empty in E/F
			if not any(ef_values):
				blank_counter += 1
				if blank_counter >= 3:
					break
				row_num += 1
				continue
			else:
				blank_counter = 0

			# Avoid appending header-like rows
			if normalize_header(str(row[1].value)).strip().upper() == "NO":
				row_num += 1
				continue

			# Collect data
			row_data = [row[i].value for i in range(1, 10)]  # B to J
			entry = [po_number, revision, supplier] + row_data
			data_rows.append(entry)
			row_num += 1

			if row_num > 500:
				log.append(f"{po_number}: Too many rows, possible error.")
				break

		return data_rows
	except Exception as e:
		log.append(f"{filepath}: Error - {str(e)}")
		return []

# --- Save Final Excel ---
def save_output(data, log_entries, output_folder):
	timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
	filename = f"Merida_Purchase_Orders_{timestamp}.xlsx"
	filepath = os.path.join(output_folder, filename)

	wb = Workbook()
	ws = wb.active
	ws.title = "Merged Orders"

	headers = ["PO Number", "Revision", "Supplier"] + STANDARD_HEADERS
	ws.append(headers)
	for row in data:
		ws.append(row)

	# Formatting
	for col in ws.columns:
		ws.column_dimensions[col[0].column_letter].auto_size = True
	ws.auto_filter.ref = ws.dimensions

	# Log sheet
	log_ws = wb.create_sheet(title="Log")
	log_ws.append(["Log Entry"])
	for entry in log_entries:
		log_ws.append([entry])

	wb.save(filepath)
	return filepath

# --- Run Script ---
def main():
	input_folder = choose_folder("Select the folder containing PO Excel files")
	output_folder = choose_folder("Select the folder to save the merged Excel file")

	files = find_excel_files(input_folder)
	all_data = []
	log = [f"Found {len(files)} Excel files"]
	
	for file in files:
		rows = extract_po_data(file, log)
		all_data.extend(rows)
		log.append(f"Processed {file} with {len(rows)} rows")

	final_path = save_output(all_data, log, output_folder)
	print(f"Merge complete. Output saved to {final_path}")

if __name__ == '__main__':
	main()
