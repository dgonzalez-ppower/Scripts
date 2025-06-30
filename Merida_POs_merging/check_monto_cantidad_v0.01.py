import os
import re
import unicodedata
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook

VALID_EXTENSIONS = ['.xls', '.xlsx', '.xlsm']

# --- GUI ---
def choose_folder(prompt):
	root = tk.Tk()
	root.withdraw()
	return filedialog.askdirectory(title=prompt)

# --- Strip accents ---
def strip_accents(text):
	return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

# --- Find Excel Files ---
def find_excel_files(root_folder):
	matched_files = []
	for dirpath, _, files in os.walk(root_folder):
		for file in files:
			if any(file.lower().endswith(ext) for ext in VALID_EXTENSIONS):
				name_clean = strip_accents(file.lower())
				if name_clean.startswith("pmx-ccc") and "entregas" not in name_clean:
					matched_files.append(os.path.join(dirpath, file))
	return matched_files

# --- Get value from merged or regular cell ---
def get_merged_value(ws, row, col):
	for merged in ws.merged_cells.ranges:
		if (row, col) in merged.cells:
			if (row, col) != (merged.min_row, merged.min_col):
				return ""  # Clean duplicates from merged ranges
			return ws.cell(row=merged.min_row, column=merged.min_col).value
	return ws.cell(row=row, column=col).value

# --- Extract data block from target file ---
def extract_data_block(filepath, header_row_idx, header_cols):
	wb = load_workbook(filepath, data_only=True)
	ws = wb.active
	data_rows = []
	blank_count = 0
	row_idx = header_row_idx + 1

	# Include original headers
	header_values = [get_merged_value(ws, header_row_idx, col) for col in header_cols]
	data_rows.append(header_values)

	while row_idx <= ws.max_row:
		row_data = []
		empty = True
		for col in header_cols:
			val = get_merged_value(ws, row_idx, col)
			if isinstance(val, str):
				val = val.strip()
			if val:
				empty = False
			row_data.append(val if val else "")
		if empty:
			blank_count += 1
			if blank_count >= 3:
				break
		else:
			blank_count = 0
			data_rows.append(row_data)
		row_idx += 1
	return data_rows

# --- Search for "monto" and "cantidad" and extract headers ---
def search_terms_in_file(filepath):
	try:
		wb = load_workbook(filepath, data_only=True)
		ws = wb.active
		monto_pos = cantidad_pos = "NOT FOUND"
		header_texts = [""] * 12
		header_cols = [0] * 12
		header_row_idx = None

		# Search for "cantidad" row
		for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
			for cell in row:
				if cell.value and isinstance(cell.value, str) and "cantidad" in strip_accents(cell.value.lower()):
					cantidad_pos = cell.coordinate
					header_row_idx = cell.row
					break
			if cantidad_pos != "NOT FOUND":
				break

		# Search for "monto"
		for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
			for cell in row:
				if cell.value and isinstance(cell.value, str) and "monto" in strip_accents(cell.value.lower()):
					monto_pos = cell.coordinate
					break
			if monto_pos != "NOT FOUND":
				break

		if header_row_idx:
			merged_ranges = [r for r in ws.merged_cells.ranges if r.min_row == header_row_idx]
			for r in merged_ranges:
				ws.unmerge_cells(str(r))

			headers_found = []
			cols_found = []
			cantidad_col_index = None

			# Scan headers
			for col in range(1, ws.max_column + 1):
				val = ws.cell(row=header_row_idx, column=col).value
				if isinstance(val, str) and val.strip():
					headers_found.append(val.strip())
					cols_found.append(col)
					if "cantidad" in strip_accents(val.lower()) and cantidad_col_index is None:
						cantidad_col_index = len(headers_found) - 1
				else:
					headers_found.append("")
					cols_found.append(col)

			# Adjust to insert cantidad at index 7 (Header Text 8)
			if cantidad_col_index is not None:
				pad_left = max(0, 7 - cantidad_col_index)
				headers_found = ([""] * pad_left) + headers_found
				cols_found = ([0] * pad_left) + cols_found
				headers_found = headers_found[:12] + [""] * (12 - len(headers_found))
				cols_found = cols_found[:12] + [0] * (12 - len(cols_found))
			else:
				headers_found = headers_found[:12] + [""] * (12 - len(headers_found))
				cols_found = cols_found[:12] + [0] * (12 - len(cols_found))

			data_preview = []
			if "PMX-CCC MÉRIDA II-107-4 Pedido" in filepath:
				data_preview = extract_data_block(filepath, header_row_idx, cols_found)

			return [os.path.basename(filepath), monto_pos, cantidad_pos], [os.path.basename(filepath)] + headers_found, data_preview

		else:
			return [os.path.basename(filepath), monto_pos, cantidad_pos], [os.path.basename(filepath)] + [""] * 12, []

	except Exception as e:
		return [os.path.basename(filepath), f"ERROR: {str(e)}", f"ERROR: {str(e)}"], [os.path.basename(filepath)] + [f"ERROR: {str(e)}"] + [""] * 11, []

# --- Save Output ---
def save_report(position_data, header_data, data_preview, output_folder):
	wb = Workbook()
	ws1 = wb.active
	ws1.title = "Header Positions"
	ws1.append(["File Name", "Posicion MONTO", "Posicion CANTIDAD"])
	for row in position_data:
		ws1.append(row)

	ws2 = wb.create_sheet(title="Extracted Headers")
	ws2.append(["File Name"] + [f"Header Text {i+1}" for i in range(12)])
	for row in header_data:
		ws2.append(row)

	if data_preview:
		ws3 = wb.create_sheet(title="Data Preview")
		for row in data_preview:
			ws3.append(["PMX-CCC MÉRIDA II-107-4 Pedido.xlsx"] + row)

	filepath = os.path.join(output_folder, "Header_Position_Report.xlsx")
	wb.save(filepath)
	print(f"Report saved to {filepath}")

# --- Main ---
def main():
	input_folder = choose_folder("Select folder with Excel files")
	output_folder = choose_folder("Select folder to save the report")
	files = find_excel_files(input_folder)
	position_results = []
	header_results = []
	data_preview_rows = []
	for file in files:
		pos_data, header_data, data_preview = search_terms_in_file(file)
		position_results.append(pos_data)
		header_results.append(header_data)
		if data_preview:
			data_preview_rows = data_preview
	save_report(position_results, header_results, data_preview_rows, output_folder)

if __name__ == '__main__':
	main()