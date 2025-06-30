import os
import re
import unicodedata
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook, Workbook

VALID_EXTENSIONS = ['.xls', '.xlsx', '.xlsm']

# --- GUI ---
def choose_folder(prompt):
	root = tk.Tk()
	root.withdraw()
	return filedialog.askdirectory(title=prompt)

# --- Strip accents ---
def strip_accents(text):
	return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

# --- Find Excel Files ---
def find_excel_files(root_folder):
	matched_files = []
	for dirpath, _, files in os.walk(root_folder):
		for file in files:
			if any(file.lower().endswith(ext) for ext in VALID_EXTENSIONS):
				name_clean = strip_accents(file.lower())
				if name_clean.startswith("pmx-ccc") and "entregas" not in name_clean:
					matched_files.append(os.path.join(dirpath, file))
	return matched_files

# --- Search for "monto" and "cantidad" and extract headers ---
def search_terms_in_file(filepath):
	try:
		wb = load_workbook(filepath, data_only=True)
		ws = wb.active
		monto_pos = cantidad_pos = "NOT FOUND"
		header_texts = []

		# Search "monto" from row 10 down
		for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
			for cell in row:
				if cell.value and isinstance(cell.value, str):
					value = strip_accents(cell.value.lower())
					if "monto" in value and monto_pos == "NOT FOUND":
						monto_pos = cell.coordinate
						break
			if monto_pos != "NOT FOUND":
				break

		# Search "cantidad" from row 10 down
		cantidad_cell = None
		for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
			for cell in row:
				if cell.value and isinstance(cell.value, str):
					value = strip_accents(cell.value.lower())
					if "cantidad" in value and cantidad_pos == "NOT FOUND":
						cantidad_pos = cell.coordinate
						cantidad_cell = cell
						break
			if cantidad_pos != "NOT FOUND":
				break

		# Reconstruct header from cantidad_cell considering merged cells
		if cantidad_cell:
			row_idx = cantidad_cell.row

			def get_merged_range_end(col_idx):
				for merged in ws.merged_cells.ranges:
					if (row_idx, col_idx) in merged.cells:
						return (ws.cell(row=merged.min_row, column=merged.min_col).value, merged.max_col)
				return (ws.cell(row=row_idx, column=col_idx).value, col_idx)

			# Build header list from left to right
			col_index = cantidad_cell.column
			header_texts = []

			# Look left
			left_headers = []
			i = col_index - 1
			while i > 0:
				val, end_col = get_merged_range_end(i)
				if isinstance(val, str) and val.strip():
					left_headers.insert(0, val.strip())
					i = i - (end_col - i + 1)
				else:
					break

			# Look right
			right_headers = []
			i = col_index + 1
			while i <= ws.max_column:
				val, end_col = get_merged_range_end(i)
				if isinstance(val, str) and val.strip():
					right_headers.append(val.strip())
					i = end_col + 1
				else:
					break

			# Compose final header list ensuring "cantidad" lands in index 5 (Header Text 6)
			headers = left_headers + [cantidad_cell.value.strip()] + right_headers
			if len(left_headers) >= 5:
				header_texts = headers[:12]
			else:
				pad = [""] * (5 - len(left_headers))
				header_texts = pad + headers
				header_texts = header_texts[:12]
			while len(header_texts) < 12:
				header_texts.append("")

		return [os.path.basename(filepath), monto_pos, cantidad_pos], [os.path.basename(filepath)] + header_texts
	except Exception as e:
		return [os.path.basename(filepath), f"ERROR: {str(e)}", f"ERROR: {str(e)}"], [os.path.basename(filepath)] + [f"ERROR: {str(e)}"] + [""] * 11

# --- Save Output ---
def save_report(position_data, header_data, output_folder):
	wb = Workbook()
	ws1 = wb.active
	ws1.title = "Header Positions"
	ws1.append(["File Name", "Posicion MONTO", "Posicion CANTIDAD"])
	for row in position_data:
		ws1.append(row)

	ws2 = wb.create_sheet(title="Extracted Headers")
	ws2.append(["File Name"] + [f"Header Text {i+1}" for i in range(12)])
	for row in header_data:
		ws2.append(row)

	filepath = os.path.join(output_folder, "Header_Position_Report.xlsx")
	wb.save(filepath)
	print(f"Report saved to {filepath}")

# --- Main ---
def main():
	input_folder = choose_folder("Select folder with Excel files")
	output_folder = choose_folder("Select folder to save the report")
	files = find_excel_files(input_folder)
	position_results = []
	header_results = []
	for file in files:
		pos_data, header_data = search_terms_in_file(file)
		position_results.append(pos_data)
		header_results.append(header_data)
	save_report(position_results, header_results, output_folder)

if __name__ == '__main__':
	main()