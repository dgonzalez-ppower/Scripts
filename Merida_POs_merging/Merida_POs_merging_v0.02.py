import os
import re
import unicodedata
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

VALID_EXTENSIONS = ['.xls', '.xlsx', '.xlsm']

def choose_folder(prompt):
    root = tk.Tk()
    root.withdraw()
    return filedialog.askdirectory(title=prompt)

def strip_accents(text):
    return ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')

def contains_reembolso(filepath):
    try:
        wb = load_workbook(filepath, data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "reembolso" in strip_accents(cell.value.lower()):
                            return True
    except:
        pass
    return False

def find_excel_files(root_folder):
    matched_files = []
    for dirpath, _, files in os.walk(root_folder):
        for file in files:
            if any(file.lower().endswith(ext) for ext in VALID_EXTENSIONS):
                name_clean = strip_accents(file.lower())
                if name_clean.startswith("pmx-ccc") and "entregas" not in name_clean:
                    matched_files.append(os.path.join(dirpath, file))
    return matched_files

def get_merged_value(ws, row, col):
    if col < 1 or row < 1:
        return ""
    for merged in ws.merged_cells.ranges:
        if (row, col) in merged.cells:
            if (row, col) != (merged.min_row, merged.min_col):
                return ""
            return ws.cell(row=merged.min_row, column=merged.min_col).value
    return ws.cell(row=row, column=col).value

def extract_revision_and_supplier(ws):
    revision = "NO DATA FOUND"
    for row in [ws[1], ws[2]]:
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "rev" in cell.value.lower():
                revision = cell.value.strip()
                break
        if revision != "NO DATA FOUND":
            break

    supplier = "NO DATA FOUND"
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=11):
        for idx, cell in enumerate(row):
            if cell.value and isinstance(cell.value, str):
                val = strip_accents(cell.value.lower())
                if "razon" in val and "social" in val:
                    for right_cell in row[idx+1:]:
                        if right_cell.value and isinstance(right_cell.value, str):
                            supplier = right_cell.value.strip()
                            break
                    break
        if supplier != "NO DATA FOUND":
            break
    return revision, supplier

def extract_header_and_data(ws):
    monto_pos = cantidad_pos = "NOT FOUND"
    header_cols = [1] * 12
    header_row_idx = None

    for row in ws.iter_rows(min_row=10, max_row=ws.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "cantidad" in strip_accents(cell.value.lower()):
                cantidad_pos = cell.coordinate
                header_row_idx = cell.row
                break
        if cantidad_pos != "NOT FOUND":
            break

    if not header_row_idx:
        return [""] * 12, [1] * 12, -1

    merged_ranges = [r for r in ws.merged_cells.ranges if r.min_row == header_row_idx]
    for r in merged_ranges:
        ws.unmerge_cells(str(r))

    headers_found = []
    cols_found = []
    cantidad_col_index = None

    for col in range(1, ws.max_column + 1):
        val = get_merged_value(ws, header_row_idx, col)
        if isinstance(val, str) and val.strip():
            headers_found.append(val.strip())
            cols_found.append(col)
            if "cantidad" in strip_accents(val.lower()) and cantidad_col_index is None:
                cantidad_col_index = len(headers_found) - 1
        else:
            headers_found.append("")
            cols_found.append(col)

    if cantidad_col_index is not None:
        pad_left = max(0, 7 - cantidad_col_index)
        headers_found = ([""] * pad_left) + headers_found
        cols_found = ([1] * pad_left) + cols_found
    headers_found = headers_found[:12] + [""] * (12 - len(headers_found))
    cols_found = cols_found[:12] + [1] * (12 - len(cols_found))
    return headers_found, cols_found, header_row_idx

def extract_data(ws, header_row_idx, header_cols):
    data_rows = []
    row_idx = header_row_idx + 1
    blank_count = 0
    while row_idx <= ws.max_row:
        row_data = []
        empty = True
        stop_by_subtotal = False
        for col in header_cols:
            val = get_merged_value(ws, row_idx, col)
            if isinstance(val, str):
                val = val.strip()
                if "subtotal" in strip_accents(val.lower()):
                    stop_by_subtotal = True
            if val:
                empty = False
            row_data.append(val if val else "")
        if stop_by_subtotal:
            break
        if empty:
            blank_count += 1
            if blank_count >= 3:
                break
        else:
            blank_count = 0
            data_rows.append(row_data)
        row_idx += 1
    return data_rows

def main():
    input_folder = choose_folder("Select folder with Excel files")
    output_folder = choose_folder("Select folder to save the merged Excel file")
    files = find_excel_files(input_folder)
    merged_data = []
    log = [f"Found {len(files)} Excel files"]

    for file in files:
        if contains_reembolso(file):
            log.append(f"{os.path.basename(file)}: SKIPPED (contains 'reembolso')")
            continue  # Salta a la siguiente iteración
        
        file_name = os.path.basename(file)
        po_number = os.path.splitext(file_name)[0]
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            revision, supplier = extract_revision_and_supplier(ws)
            headers, cols, header_row_idx = extract_header_and_data(ws)
            if header_row_idx == -1:
                raise ValueError("Header row with 'cantidad' not found.")
            data_rows = extract_data(ws, header_row_idx, cols)
            if not data_rows:
                data_rows = [["NO DATA FOUND"] * 12]
            for row in data_rows:
                merged_data.append([po_number, revision, supplier] + row)
            log.append(f"{file_name}: {len(data_rows)} rows added.")
        except Exception as e:
            merged_data.append([po_number, "NO DATA FOUND", "NO DATA FOUND"] + ["NO DATA FOUND"] * 12)
            log.append(f"{file_name}: ERROR - {str(e)}")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Merged Orders"
    headers_out = ["PO Number", "Revision", "Supplier"] + [f"Header Text {i+1}" for i in range(12)]
    ws_out.append(headers_out)
    for row in merged_data:
        ws_out.append(row)
    ws_out.auto_filter.ref = f"A1:{get_column_letter(len(headers_out))}1"

    log_ws = wb_out.create_sheet(title="Log")
    log_ws.append(["Log Entry"])
    for entry in log:
        log_ws.append([entry])

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_path = os.path.join(output_folder, f"Merida_Purchase_Orders_{timestamp}.xlsx")
    wb_out.save(output_path)
    print(f"Saved to: {output_path}")

if __name__ == "__main__":
    main()